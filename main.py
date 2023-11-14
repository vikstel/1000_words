import json
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests
import lxml
import csv



url = 'https://www.kreekly.com/lists/1000-naibolee-populyarnyh-angliyskih-slov/'
headers = {
    "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q="
              "0.8,application/signed-exchange;v=b3;q=0.7",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 "
                  "Safari/537.36"
}
response = requests.get(url, headers=headers, verify=False)
soup = BeautifulSoup(response.text, 'lxml')
all_words = soup.find("div", class_="list-words")
all_words_list = []
count = 0
for words in all_words:
    english_word = words.find("span", class_="eng").text
    russian_word = words.find("span", class_="rus").text
    img_link = f'https://www.kreekly.com{words.find("img").get("src")}'
    all_words_list.append(
        [english_word, russian_word, img_link]
    )
    count += 1
    print(english_word)
    print(russian_word)
    print(img_link)
    print("-" * 30)
    print(count)
with open('words.json', 'w', encoding='utf-8') as file:
    json.dump(all_words_list, file, indent=4, ensure_ascii=False)


# wb = load_workbook('./1000_words.xlsx')
# ws = wb['Лист1']
# words_dict = {}
# row_a = ws['A1': 'B2496']
# for row in row_a:
#     words_dict[row[0].value] = row[1].value
# print(words_dict)